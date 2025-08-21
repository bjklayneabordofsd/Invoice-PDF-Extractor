import pdfplumber
import re
import openpyxl
from datetime import datetime
from extract_invoice_data import extract_invoice_data

def add_to_excel(invoice_data, excel_file="Invoice_Spreadsheet.xlsx", pdf_file="Oaks Invoices 1- 10 .pdf"):
    """Add invoice data to Excel file and log to History sheet"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        # Get the first sheet (main invoice sheet) - NOT the active sheet
        invoice_sheet = wb.worksheets[0]  # First sheet should be the invoice data sheet
        
        if isinstance(invoice_data, dict):
            invoice_data = [invoice_data]
        
        # Find next empty row in main invoice sheet
        next_row = 2
        while invoice_sheet.cell(row=next_row, column=1).value:
            next_row += 1
        
        # Add data to main invoice sheet
        records_added = 0
        for invoice in invoice_data:
            # Add all rows, replace "skip" with empty string
            invoice_sheet.cell(row=next_row, column=1, value='Oaks at Creekside')  # Property Name (constant)
            
            # For each field, replace "skip" with empty string
            vendor = invoice.get('vendor_name', '')
            invoice_sheet.cell(row=next_row, column=2, value='' if vendor == 'skip' else vendor)
            
            service = invoice.get('service_type', '')
            invoice_sheet.cell(row=next_row, column=3, value='' if service == 'skip' else service)
            
            invoice_num = invoice.get('invoice_number', '')
            invoice_sheet.cell(row=next_row, column=4, value='' if invoice_num == 'skip' else invoice_num)
            
            due_date = invoice.get('invoice_date', '')
            invoice_sheet.cell(row=next_row, column=5, value='' if due_date == 'skip' else due_date)
            
            amount = invoice.get('invoice_amount', '')
            invoice_sheet.cell(row=next_row, column=6, value='' if amount == 'skip' else amount)
            
            next_row += 1
            records_added += 1
        
        # Add entry to History sheet (second sheet)
        if 'History' in wb.sheetnames:
            history_sheet = wb['History']
            
            # Find next empty row in History sheet
            history_row = 2
            while history_sheet.cell(row=history_row, column=1).value:
                history_row += 1
            
            # Add history entry
            current_date = datetime.now().strftime('%m/%d/%Y')
            current_time = datetime.now().strftime('%I:%M %p')
            
            history_sheet.cell(row=history_row, column=1, value=current_date)  # Date
            history_sheet.cell(row=history_row, column=2, value=current_time)  # Time
            history_sheet.cell(row=history_row, column=3, value='')  # Version (manual)
            history_sheet.cell(row=history_row, column=4, value='')  # Author (manual)
            history_sheet.cell(row=history_row, column=5, value=f'Processed {pdf_file}')  # Description
            history_sheet.cell(row=history_row, column=6, value=records_added)  # Records Added
        
        wb.save(excel_file)
        wb.close()
        print(f"✓ Added {records_added} invoice(s) to main sheet")
        print(f"✓ Logged to History sheet")
            
    except PermissionError:
        backup_name = f"Invoice_Backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(backup_name)
        wb.close()
        print(f"⚠️ Original file is open. Saved as: {backup_name}")
    except FileNotFoundError:
        print(f"❌ Excel file '{excel_file}' not found. Please make sure the file exists.")

def handle_multi_page_invoices(pdf_file="Oaks Invoices 1- 10 .pdf"):
    """
    Step 1: Combine multi-page invoices first
    """
    combined_pages = []
    combined_page_list = []
    
    with pdfplumber.open(pdf_file) as pdf:
        i = 0
        while i < len(pdf.pages):
            text = pdf.pages[i].extract_text() or ""
            
            # Check if this is page 1 of 2
            if re.search(r'Page 1 of 2', text, re.I):
                # Combine with next page
                if i + 1 < len(pdf.pages):
                    text2 = pdf.pages[i + 1].extract_text() or ""
                    combined_text = text + "\n" + text2
                    combined_pages.append({
                        'page_nums': [i, i+1],
                        'text': combined_text
                    })
                    combined_page_list.append(f"{i+1} and {i+2}")
                    i += 2  
                    continue
            
            # Single page
            combined_pages.append({
                'page_nums': [i],
                'text': text
            })
            i += 1
    
    # Store combined pages info for summary
    combined_pages_info = combined_page_list
    return combined_pages, combined_pages_info

def identify_invoice_pages(combined_pages):
    """
    Step 2: Identify which pages/combined pages are invoices
    """
    invoice_pages = []
    non_invoice_pages = []
    
    for idx, page_data in enumerate(combined_pages):
        text = page_data['text']
        
        # Three simple checks
        has_invoice_word = bool(re.search(r'invoice', text, re.I))
        has_date = bool(re.search(r'\d{1,2}/\d{1,2}/\d{4}', text))
        has_money = bool(re.search(r'[\$\s][\d,]+\.\d{2}', text))
        
        if has_invoice_word and has_date and has_money:
            invoice_pages.append(page_data)
        else:
            non_invoice_pages.append(page_data)
    
    return invoice_pages, len(invoice_pages), len(non_invoice_pages)


def extract_from_pdf(pdf_file): 
    """
    Complete extraction pipeline
    """
    # Step 1: Handle multi-page
    combined, combined_info = handle_multi_page_invoices(pdf_file)
    
    # Step 2: Identify invoices
    invoices, valid_count, invalid_count = identify_invoice_pages(combined)
    

    print("\nSummary:")
    print(f"  - Valid invoice pages: {valid_count}")
    print(f"  - Non-invoice pages: {invalid_count}")
    if combined_info:
        for pages in combined_info:
            print(f"  - Combined pages {pages}")
    
    # Step 3: Extract data
    data = extract_invoice_data(invoices)
    
    return data


if __name__ == "__main__":
    # change the file name "Oaks Invoices 1- 10 .pdf" if there is another file
    pdf_filename = "Oaks Invoices 1- 10 .pdf"
    
    # Extract data from PDF
    data = extract_from_pdf(pdf_filename)

    print("\n📊 FINAL SUMMARY TABLE:")
    print("=" * 120)
    print(f"{'#':<3} {'Vendor':<25} {'Service Type':<40} {'Invoice #':<12} {'Due Date':<12} {'Amount':<10}")
    print("-" * 120)
    
    for i, entry in enumerate(data, 1):
        vendor = entry['vendor_name'][:25]
        service = entry.get('service_type', 'skip')[:40]
        invoice_num = entry.get('invoice_number', 'skip')
        due_date = entry.get('invoice_date', 'skip')
        amount = entry.get('invoice_amount', 'skip')
        
        print(f"{i:<3} {vendor:<25} {service:<40} {invoice_num:<12} {due_date:<12} {amount:<10}")
    
    print("-" * 120)
    

    print("\n📝 Adding data to Excel...")
    add_to_excel(data, pdf_file=pdf_filename)